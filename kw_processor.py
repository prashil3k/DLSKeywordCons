#!/usr/bin/env python3
"""
Zero-Volume Keyword Processor for Demo-Led SEO
================================================
Takes raw keyword exports and:
1. Categorizes (integrations, how-tos, tutorials, irrelevant)
2. Semantically deduplicates (groups variants of the same intent)
3. Ranks by signal strength (more variants = more demand)
4. Outputs clean, actionable keyword list with canonical titles

Usage:
    python3 kw_processor.py <input_file> <tool_name> [--output <output_file>]

Input: .txt, .csv, .xlsx, or .xls file (reads first column for keywords)
Output: CSV with columns: canonical_title, category, variant_count, signal_strength, variants
"""

import argparse
import csv
import re
import sys
from collections import defaultdict
from pathlib import Path


def load_keywords(filepath, column=None):
    """Load keywords from .txt, .csv, .xlsx, or .xls file.

    For Excel/CSV: reads the specified column (name or index), or auto-detects
    the first column that looks like keywords. Skips header rows automatically.

    Args:
        filepath: path to input file
        column: column name or 0-based index to read from (optional, auto-detects)
    """
    filepath = str(filepath)
    ext = Path(filepath).suffix.lower()

    # Common header names to skip
    skip_values = {'name', 'keyword', 'keywords', 'query', 'queries',
                   'search term', 'search terms', 'term', 'terms',
                   'topic', 'title', 'url', ''}

    if ext in ('.xlsx', '.xlsm'):
        return _load_from_xlsx(filepath, column, skip_values)
    elif ext == '.xls':
        return _load_from_xls(filepath, column, skip_values)
    elif ext == '.csv':
        return _load_from_csv(filepath, column, skip_values)
    else:
        # Plain text fallback
        return _load_from_txt(filepath, skip_values)


def _load_from_xlsx(filepath, column, skip_values):
    """Load keywords from .xlsx/.xlsm using openpyxl."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    keywords = []

    col_idx = _resolve_column_xlsx(ws, column)

    for row in ws.iter_rows(min_col=col_idx + 1, max_col=col_idx + 1):
        cell = row[0]
        if cell.value is not None:
            val = str(cell.value).strip()
            if val.lower() not in skip_values:
                keywords.append(val)

    wb.close()
    return keywords


def _resolve_column_xlsx(ws, column):
    """Find the right column index in an xlsx worksheet."""
    if column is not None:
        if isinstance(column, int):
            return column
        # Find column by header name
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value and str(cell.value).strip().lower() == column.lower():
                return cell.column - 1
    return 0  # Default to first column


def _load_from_xls(filepath, column, skip_values):
    """Load keywords from .xls using xlrd."""
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    keywords = []

    col_idx = 0
    if column is not None:
        if isinstance(column, int):
            col_idx = column
        else:
            # Find by header name
            for c in range(ws.ncols):
                if ws.cell_value(0, c).strip().lower() == column.lower():
                    col_idx = c
                    break

    for r in range(ws.nrows):
        val = str(ws.cell_value(r, col_idx)).strip()
        if val.lower() not in skip_values:
            keywords.append(val)

    return keywords


def _detect_encoding(filepath):
    """Detect file encoding by checking BOM bytes."""
    with open(filepath, 'rb') as f:
        raw = f.read(4)
    if raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
        return 'utf-16'
    if raw[:3] == b'\xef\xbb\xbf':
        return 'utf-8-sig'
    return 'utf-8'


def _load_from_csv(filepath, column, skip_values):
    """Load keywords from CSV file."""
    encoding = _detect_encoding(filepath)
    keywords = []
    with open(filepath, 'r', encoding=encoding) as f:
        # Sniff delimiter
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        except csv.Error:
            dialect = 'excel'

        reader = csv.reader(f, dialect)
        header = next(reader, None)

        # Auto-detect keyword column in multi-column exports (Ahrefs, Semrush, etc.)
        keyword_headers = {'keyword', 'keywords', 'query', 'queries',
                           'search term', 'search terms', 'term', 'topic', 'name'}

        col_idx = 0
        is_multi_column = header and len(header) > 1

        if header and column is not None:
            if isinstance(column, int):
                col_idx = column
            else:
                for i, h in enumerate(header):
                    if h.strip().lower() == column.lower():
                        col_idx = i
                        break
        elif header and is_multi_column:
            # Multi-column file: find the keyword column by header name
            found = False
            for i, h in enumerate(header):
                if h.strip().strip('"').lower() in keyword_headers:
                    col_idx = i
                    found = True
                    break
            if not found:
                col_idx = 0  # Fallback to first column
        elif header:
            # Single-column file: check if header row is actually data
            if header[0].strip().lower() not in skip_values:
                keywords.append(header[0].strip())

        for row in reader:
            if row and len(row) > col_idx:
                val = row[col_idx].strip()
                if val and val.lower() not in skip_values and val.lower() not in keyword_headers:
                    keywords.append(val)

    return keywords


def _load_from_txt(filepath, skip_values):
    """Load keywords from plain text file (one per line)."""
    encoding = _detect_encoding(filepath)
    keywords = []
    with open(filepath, 'r', encoding=encoding) as f:
        for line in f:
            kw = line.strip()
            if kw and kw.lower() not in skip_values:
                keywords.append(kw)
    return keywords


def normalize_tool_name(tool_name):
    """Create variations of tool name for matching."""
    name = tool_name.lower().strip()
    variants = {name}
    # Handle "software cloud" variants (Jira-specific but common pattern)
    variants.add(f"{name} software cloud")
    variants.add(f"{name} cloud")
    variants.add(f"{name} software server")
    variants.add(f"{name} software")
    variants.add(f"{name} service desk")
    variants.add(f"{name} service management")
    return variants


def categorize_keyword(kw, tool_name):
    """Categorize a keyword into: integration, howto, tutorial, or other."""
    kw_lower = kw.lower()

    # Integration keywords
    integration_signals = ['integration', 'integrate', 'connect ', 'connect to',
                           'connect with', 'sync ', 'synced']
    if any(sig in kw_lower for sig in integration_signals):
        return 'integration'

    # How-to keywords
    if kw_lower.startswith('how to') or kw_lower.startswith('how do'):
        return 'howto'

    # Tutorial keywords
    if 'tutorial' in kw_lower:
        return 'tutorial'

    # Remaining
    return 'other'


def extract_integration_pair(kw, tool_name):
    """Extract the other tool being integrated with, return a normalized key."""
    kw_lower = kw.lower()

    # Known tool names to detect (add more as needed)
    known_tools = [
        'salesforce', 'github', 'gitlab', 'slack', 'trello', 'asana', 'notion',
        'hubspot', 'zendesk', 'servicenow', 'service now', 'confluence', 'bitbucket',
        'azure devops', 'monday', 'monday.com', 'smartsheet', 'basecamp', 'clickup',
        'google sheets', 'google calendar', 'google forms', 'google docs', 'google drive',
        'gmail', 'outlook', 'teams', 'microsoft teams', 'sharepoint', 'excel',
        'microsoft excel', 'power bi', 'power automate', 'dynamics 365',
        'microsoft dynamics', 'microsoft project', 'ms project', 'ms teams',
        'datadog', 'pagerduty', 'miro', 'figma', 'figjam', 'zapier', 'make',
        'n8n', 'workato', 'airtable', 'typeform', 'surveymonkey', 'freshdesk',
        'freshservice', 'intercom', 'discord', 'whatsapp', 'tableau', 'looker',
        'snowflake', 'bigquery', 'redshift', 'postgresql', 'mysql', 'okta', 'auth0',
        'jenkins', 'circleci', 'bamboo', 'teamcity', 'terraform', 'docker',
        'kubernetes', 'aws', 'azure', 'gcp', 'sentry', 'splunk', 'new relic',
        'grafana', 'datadog', 'opsgenie', 'xray', 'zephyr', 'testrail', 'qtest',
        'sonarqube', 'snyk', 'wiz', 'crowdstrike', 'qualys', 'dynatrace',
        'workfront', 'todoist', 'coda', 'linear', 'shortcut', 'productboard',
        'aha', 'rally', 'harvest', 'toggl', 'clockify', 'tempo',
        'chatgpt', 'claude', 'gemini', 'copilot', 'cursor', 'devin',
        'pipedrive', 'copper', 'insightly', 'netsuite', 'quickbooks', 'xero',
        'stripe', 'chargebee', 'recurly', 'mailchimp', 'sendgrid', 'brevo',
        'active directory', 'ldap', 'saml', 'oauth', 'sso',
        'mcp', 'webhook', 'webhooks', 'rest api', 'api',
        'selenium', 'cypress', 'playwright', 'cucumber',
        'docusign', 'pandadoc', 'firebase', 'supabase',
        'hootsuite', 'buffer', 'semrush', 'ahrefs',
        'hp alm', 'alm', 'tfs', 'svn', 'perforce', 'git',
        'mural', 'lucidchart', 'lucidspark', 'lucid',
        'dropbox', 'box', 'onedrive', 'google workspace',
        'workday', 'sap', 'oracle', 'bamboohr', 'gusto',
        'gainsight', 'pendo', 'amplitude', 'mixpanel', 'segment',
        'postman', 'swagger', 'retool', 'appsheet', 'bubble',
    ]
    known_tools.sort(key=len, reverse=True)

    # Remove tool name variants
    tool_variants = normalize_tool_name(tool_name)
    cleaned = kw_lower
    for variant in sorted(tool_variants, key=len, reverse=True):
        cleaned = cleaned.replace(variant, ' ')

    # Try to find a known tool in the cleaned string
    for known in known_tools:
        if known in cleaned:
            return known

    # Fallback: extract by removing integration words
    remove_words = [
        'integration', 'integrate', 'integrated', 'integrating',
        'connect', 'connected', 'connecting', 'connection',
        'sync', 'synced', 'syncing',
        'how to', 'how do i', 'does', 'can', 'can we',
        'with', 'to', 'from', 'and', 'for', 'the', 'a', 'an', 'in', 'on',
        'using', 'via', 'through', 'between', 'not working', 'best', 'free',
        'official', 'features', 'documentation', 'guide', 'setup', 'steps',
        'how', 'it', 'works', 'news', 'webhook', 'rest api',
        'marketplace', 'app', 'plugin', 'tools', 'tool', 'api',
        'solutions', 'capabilities', 'services', 'service', 'plus', 'apps',
        'automation', 'diagram', 'docs', 'pricing', 'cost', 'review',
        'use cases', 'spoke', 'hub', 'enterprise', 'server', 'cloud',
        'data center', 'software', 'management', 'desk',
        '2025', '2026', 'or', 'pdf',
    ]
    for word in sorted(remove_words, key=len, reverse=True):
        cleaned = cleaned.replace(word, ' ')

    words = [w.strip().strip('?.,!') for w in cleaned.split() if w.strip()]
    words = [w for w in words if len(w) > 1 and w not in {'of', 'at', 'by', 'if', 'is', 'what', 'which'}]

    other_tool = ' '.join(words).strip()
    if not other_tool:
        other_tool = '_general_'

    return other_tool


def normalize_howto(kw, tool_name):
    """Normalize a how-to keyword to find semantic duplicates."""
    kw_lower = kw.lower().strip()

    # Remove how-to prefix
    for prefix in ['how to ', 'how do i ', 'how do you ', 'how can i ', 'how can you ']:
        if kw_lower.startswith(prefix):
            kw_lower = kw_lower[len(prefix):]
            break

    # Remove tool name variants
    tool_variants = normalize_tool_name(tool_name)
    for variant in sorted(tool_variants, key=len, reverse=True):
        kw_lower = kw_lower.replace(f' in {variant}', '')
        kw_lower = kw_lower.replace(f' on {variant}', '')
        kw_lower = kw_lower.replace(f' for {variant}', '')
        kw_lower = kw_lower.replace(f' from {variant}', '')
        kw_lower = kw_lower.replace(f' to {variant}', '')
        kw_lower = kw_lower.replace(f' with {variant}', '')
        kw_lower = kw_lower.replace(f' using {variant}', '')
        kw_lower = kw_lower.replace(variant, '')

    # Remove articles
    kw_lower = re.sub(r'\b(a|an|the|my|your|our)\b', ' ', kw_lower)

    # Remove trailing question marks and common suffixes
    kw_lower = kw_lower.rstrip('?').strip()

    # Normalize whitespace
    kw_lower = ' '.join(kw_lower.split())

    # Simple singular/plural normalization
    # "create sprints" -> "create sprint", "create labels" -> "create label"
    words = kw_lower.split()
    if len(words) >= 2:
        last = words[-1]
        if last.endswith('s') and not last.endswith('ss') and not last.endswith('us') and len(last) > 3:
            words[-1] = last[:-1]
    kw_lower = ' '.join(words)

    return kw_lower


def is_irrelevant(kw, tool_name):
    """Check if a keyword is irrelevant/low-quality for tutorial content."""
    kw_lower = kw.lower()

    irrelevant_patterns = [
        # Pronunciation/spelling
        'how to pronounce', 'how to spell', 'how to say ',
        # Meta/non-actionable
        'tutorial for beginners pdf', 'tutorial pdf', 'tutorial video',
        'tutorial youtube', 'tutorial for beginners youtube',
        'tutorial 2018', 'tutorial 2019', 'tutorial 2020', 'tutorial 2021',
        'tutorial 2022', 'tutorial 2023', 'tutorial 2024',
        'tutorial w3schools', 'tutorial udemy', 'tutorial guru99',
        'tutorial javatpoint', 'tutorial point', 'tutorial ppt',
        'tutorial powerpoint', 'tutorial hindi', 'tutorial arabic',
        'tutorial deutsch', 'tutorial español', 'tutorial français',
        'tutorial italiano', 'tutorial telugu', 'tutorial pl',
        'best tutorial', 'free tutorial', 'tutorial free',
        'tutorial for beginners video',
        # Non-tutorial queries
        'how to learn ', 'certification', 'interview', 'resume',
        'how to mention', 'what is ', 'what does ',
        'does ', ' vs ',  # comparison queries
        'how to contact', 'how to cancel subscription',
        # Too vague
        f'{tool_name.lower()} how to',
        f'how to use {tool_name.lower()}',
        f'{tool_name.lower()} tutorial',
    ]

    # Check exact matches for very generic terms
    exact_irrelevant = [
        f'{tool_name.lower()} tutorial',
        f'how to use {tool_name.lower()}',
        f'{tool_name.lower()} how to',
        f'how to {tool_name.lower()}',
        f'tutorial {tool_name.lower()}',
        f'{tool_name.lower()} tutorial for beginners',
        f'learn how to use {tool_name.lower()}',
        f'best {tool_name.lower()} tutorial',
    ]

    if kw_lower.strip() in exact_irrelevant:
        return True

    for pattern in irrelevant_patterns:
        if pattern in kw_lower:
            return True

    return False


def pick_canonical(variants, tool_name):
    """Pick the best canonical version from a group of semantic variants."""
    # Prefer "How to X in [Tool]" format
    howtos = [v for v in variants if v.lower().startswith('how to')]

    if howtos:
        # Prefer ones with articles ("a", "an") - they read more naturally
        with_article = [v for v in howtos if ' a ' in v.lower() or ' an ' in v.lower()]
        if with_article:
            # Pick the shortest natural one
            return min(with_article, key=len)
        return min(howtos, key=len)

    # For integrations, prefer "[Tool A] [Tool B] Integration" format
    integrations = [v for v in variants if 'integration' in v.lower()]
    if integrations:
        # Prefer format: "Tool A Tool B Integration"
        return min(integrations, key=len)

    return min(variants, key=len)


def generate_tutorial_title(canonical, tool_name, category):
    """Convert a canonical keyword into a proper tutorial title."""
    title = canonical.strip()

    # Already a good how-to title
    if title.lower().startswith('how to'):
        # Capitalize properly
        words = title.split()
        words[0] = 'How'
        # Capitalize tool name if present
        title = ' '.join(words)
        # Ensure tool name is properly cased
        title = re.sub(
            re.escape(tool_name.lower()),
            tool_name,
            title,
            flags=re.IGNORECASE
        )
        return title

    # Integration keyword -> How to title
    if category == 'integration':
        return title  # Keep as-is for integrations, they become their own format

    return title


def process_keywords(keywords, tool_name, verbose=False):
    """Main processing pipeline."""

    # Step 1: Categorize and filter
    categorized = {'integration': [], 'howto': [], 'tutorial': [], 'other': []}
    irrelevant = []

    for kw in keywords:
        if is_irrelevant(kw, tool_name):
            irrelevant.append(kw)
            continue
        cat = categorize_keyword(kw, tool_name)
        categorized[cat].append(kw)

    # Step 2: Deduplicate integrations
    integration_groups = defaultdict(list)
    for kw in categorized['integration']:
        key = extract_integration_pair(kw, tool_name)
        integration_groups[key].append(kw)

    # Step 3: Deduplicate how-tos
    howto_groups = defaultdict(list)
    for kw in categorized['howto']:
        key = normalize_howto(kw, tool_name)
        howto_groups[key].append(kw)

    # Step 4: Build output
    results = []

    # Process integrations
    for key, variants in sorted(integration_groups.items(), key=lambda x: -len(x[1])):
        canonical = pick_canonical(variants, tool_name)
        signal = 'high' if len(variants) >= 5 else 'medium' if len(variants) >= 3 else 'low'
        results.append({
            'canonical': canonical,
            'category': 'integration',
            'variant_count': len(variants),
            'signal': signal,
            'normalized_key': key,
            'variants': variants,
        })

    # Process how-tos
    for key, variants in sorted(howto_groups.items(), key=lambda x: -len(x[1])):
        canonical = pick_canonical(variants, tool_name)
        title = generate_tutorial_title(canonical, tool_name, 'howto')
        signal = 'high' if len(variants) >= 5 else 'medium' if len(variants) >= 3 else 'low'
        results.append({
            'canonical': title,
            'category': 'howto',
            'variant_count': len(variants),
            'signal': signal,
            'normalized_key': key,
            'variants': variants,
        })

    # Process tutorials (usually generic, mostly irrelevant)
    tutorial_group = defaultdict(list)
    for kw in categorized['tutorial']:
        # Group by rough topic
        kw_clean = kw.lower().replace(tool_name.lower(), '').replace('tutorial', '').strip()
        kw_clean = ' '.join(kw_clean.split())
        if kw_clean:
            tutorial_group[kw_clean].append(kw)
        else:
            irrelevant.append(kw)

    for key, variants in sorted(tutorial_group.items(), key=lambda x: -len(x[1])):
        canonical = pick_canonical(variants, tool_name)
        signal = 'high' if len(variants) >= 5 else 'medium' if len(variants) >= 3 else 'low'
        results.append({
            'canonical': canonical,
            'category': 'tutorial',
            'variant_count': len(variants),
            'signal': signal,
            'normalized_key': key,
            'variants': variants,
        })

    # Process other
    for kw in categorized['other']:
        results.append({
            'canonical': kw,
            'category': 'other',
            'variant_count': 1,
            'signal': 'low',
            'normalized_key': kw.lower(),
            'variants': [kw],
        })

    return results, irrelevant, categorized


def cross_reference_existing(results, existing_titles, tool_name):
    """Mark which processed keywords already have articles."""
    tool_lower = tool_name.lower()
    # Normalize existing titles for matching
    existing_normalized = set()
    for title in existing_titles:
        if tool_lower in title.lower():
            normalized = title.lower().strip()
            existing_normalized.add(normalized)

    for result in results:
        canonical_lower = result['canonical'].lower().strip()
        # Check if any variant matches existing titles
        matched = canonical_lower in existing_normalized
        if not matched:
            for variant in result['variants']:
                if variant.lower().strip() in existing_normalized:
                    matched = True
                    break
        result['already_covered'] = matched

    return results


def output_results(results, irrelevant, categorized, tool_name, output_path=None):
    """Output processed results."""

    # Stats
    total_input = sum(len(v) for v in categorized.values()) + len(irrelevant)
    total_integrations = len([r for r in results if r['category'] == 'integration'])
    total_howtos = len([r for r in results if r['category'] == 'howto'])
    total_tutorials = len([r for r in results if r['category'] == 'tutorial'])
    total_other = len([r for r in results if r['category'] == 'other'])
    new_only = [r for r in results if not r.get('already_covered', False)]

    print(f"\n{'='*80}")
    print(f"  KEYWORD PROCESSING RESULTS — {tool_name}")
    print(f"{'='*80}")
    print(f"\n  Input keywords:        {total_input}")
    print(f"  Irrelevant/filtered:   {len(irrelevant)}")
    print(f"  Unique intents found:  {len(results)}")
    if any('already_covered' in r for r in results):
        print(f"  Already covered:       {len(results) - len(new_only)}")
        print(f"  NEW opportunities:     {len(new_only)}")
    print(f"\n  By category:")
    print(f"    Integrations:  {total_integrations}")
    print(f"    How-tos:       {total_howtos}")
    print(f"    Tutorials:     {total_tutorials}")
    print(f"    Other:         {total_other}")

    # High-signal opportunities
    high_signal = sorted(
        [r for r in results if r['signal'] in ('high', 'medium') and not r.get('already_covered', False)],
        key=lambda x: -x['variant_count']
    )

    if high_signal:
        print(f"\n{'─'*80}")
        print(f"  TOP NEW OPPORTUNITIES (by variant count = demand signal)")
        print(f"{'─'*80}")

        for i, r in enumerate(high_signal[:50], 1):
            marker = '★' if r['signal'] == 'high' else '●'
            print(f"\n  {marker} [{r['variant_count']} variants] {r['canonical']}")
            print(f"    Category: {r['category']} | Signal: {r['signal']}")
            if len(r['variants']) <= 5:
                for v in r['variants']:
                    print(f"      - {v}")
            else:
                for v in r['variants'][:3]:
                    print(f"      - {v}")
                print(f"      ... and {len(r['variants'])-3} more")

    # Write CSV output
    if output_path:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow([
                'canonical_keyword', 'suggested_title', 'category',
                'variant_count', 'signal_strength', 'already_covered',
                'all_variants'
            ])
            for r in sorted(results, key=lambda x: (-x['variant_count'], x['category'])):
                # Generate suggested title
                if r['category'] == 'howto':
                    suggested = r['canonical']
                elif r['category'] == 'integration':
                    suggested = f"How to Integrate {tool_name} with {r['normalized_key'].title()}"
                else:
                    suggested = r['canonical']

                writer.writerow([
                    r['canonical'],
                    suggested,
                    r['category'],
                    r['variant_count'],
                    r['signal'],
                    r.get('already_covered', ''),
                    ' | '.join(r['variants'][:10]),
                ])
        print(f"\n  ✓ Full results written to: {output_path}")

    # Also write irrelevant for review
    if output_path and irrelevant:
        irrelevant_path = output_path.replace('.csv', '_irrelevant.csv')
        with open(irrelevant_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['keyword', 'reason'])
            for kw in irrelevant:
                writer.writerow([kw, 'filtered'])
        print(f"  ✓ Filtered keywords written to: {irrelevant_path}")

    print()


def load_all_existing(published_file=None, writer_folder=None, published_col=None):
    """Load all 'already covered' keywords from published CSV + writer sheets folder.

    Returns a flat list of all keyword/title strings.
    """
    all_existing = []

    # Load published articles
    if published_file:
        titles = load_keywords(published_file, column=published_col)
        print(f"  Published articles: {len(titles)} from {Path(published_file).name}")
        all_existing.extend(titles)

    # Load all writer sheets (all tabs, first column)
    if writer_folder:
        import openpyxl
        folder = Path(writer_folder)
        skip_values = {'name', 'keyword', 'keywords', 'query', 'queries',
                       'search term', 'search terms', 'term', 'terms',
                       'topic', 'title', 'url', ''}

        writer_count = 0
        for fpath in sorted(folder.glob('*.xls*')):
            try:
                wb = openpyxl.load_workbook(str(fpath), read_only=True, data_only=True)
                file_count = 0
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                        if row[0] is not None:
                            val = str(row[0]).strip()
                            if val and val.lower() not in skip_values:
                                # Skip section headers (bold markers like "**Cisco Slido**")
                                if val.startswith('**') and val.endswith('**'):
                                    continue
                                all_existing.append(val)
                                file_count += 1
                wb.close()
                writer_count += file_count
                print(f"  Writer sheet: {fpath.name} → {file_count} keywords across {len(wb.sheetnames)} tabs")
            except Exception as e:
                print(f"  WARNING: Could not read {fpath.name}: {e}")

        print(f"  Writer sheets total: {writer_count}")

    print(f"  Combined reference set: {len(all_existing)} keywords/titles")
    return all_existing


def main():
    parser = argparse.ArgumentParser(description='Process zero-volume keywords for demo-led SEO')
    parser.add_argument('input_file', help='Input file (.txt, .csv, .xlsx, .xls)')
    parser.add_argument('tool_name', help='Name of the SaaS tool (e.g., "Jira", "Salesforce")')
    parser.add_argument('--output', '-o', help='Output CSV file path')
    parser.add_argument('--column', '-c', help='Column name or index to read keywords from (default: first column)')
    parser.add_argument('--existing', '-e', help='File with already-published article titles (.txt, .csv, .xlsx, .xls)')
    parser.add_argument('--existing-column', help='Column name/index for existing titles file (default: first column)')
    parser.add_argument('--writers', '-w', help='Folder containing writer XLS sheets (reads all tabs, first column)')
    parser.add_argument('--verbose', '-v', action='store_true')

    args = parser.parse_args()

    # Parse column argument (could be int or string)
    col = args.column
    if col is not None:
        try:
            col = int(col)
        except ValueError:
            pass  # Keep as string (column name)

    existing_col = args.existing_column
    if existing_col is not None:
        try:
            existing_col = int(existing_col)
        except ValueError:
            pass

    # Load keywords
    keywords = load_keywords(args.input_file, column=col)
    print(f"Loaded {len(keywords)} keywords from {args.input_file}")

    # Process
    results, irrelevant, categorized = process_keywords(keywords, args.tool_name, args.verbose)

    # Cross-reference with existing articles if provided
    if args.existing or args.writers:
        print(f"\nLoading reference set (already covered):")
        existing = load_all_existing(
            published_file=args.existing,
            writer_folder=args.writers,
            published_col=existing_col,
        )
        results = cross_reference_existing(results, existing, args.tool_name)

    # Default output path
    output_path = args.output
    if not output_path:
        stem = Path(args.input_file).stem
        output_path = str(Path(args.input_file).parent / f"{stem}_processed.csv")

    # Output
    output_results(results, irrelevant, categorized, args.tool_name, output_path)


if __name__ == '__main__':
    main()
