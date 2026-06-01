#!/usr/bin/env python3
"""
Keyword Consolidation Pipeline
===============================
Reads raw keyword XLSX (multi-tab), deduplicates, normalizes, filters against
the CMS export, and outputs clean per-tool CSVs.

Usage:
    python3 pipeline.py --xlsx "demo led keywords - WIP.xlsx" --cms "Storylane - Tutorials - *.csv"
    python3 pipeline.py --xlsx input.xlsx --cms cms.csv --tools "capcut,docker"
    python3 pipeline.py --xlsx input.xlsx --cms cms.csv --ai-cleanup
"""

import argparse
import csv
import os
import re
import sys
import json
import time
from collections import defaultdict, Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ── Constants ────────────────────────────────────────────────────────────────

OUTPUT_DIR = Path("output")

SKIP_TABS = {
    "dump", "clean", "all first pass", "tool closed in phase 1 (ignore",
    "paypal (ignore)",
}

STOP_WORDS = {
    'a', 'an', 'the', 'in', 'on', 'to', 'for', 'from', 'with', 'of', 'at',
    'by', 'and', 'or', 'is', 'it', 'my', 'your', 'our', 'how', 'do', 'i',
    'you', 'can', 'does', 'using', 'via', 'through', 'into', 'be', 'been',
    'being', 'was', 'were', 'are', 'am', 'has', 'have', 'had', 'will',
    'would', 'could', 'should', 'shall', 'may', 'might', 'must',
    'that', 'this', 'these', 'those', 'which', 'what', 'where', 'when',
    'who', 'whom', 'whose', 'why', 'so', 'if', 'then', 'than',
    'but', 'not', 'no', 'nor', 'just', 'also', 'too', 'very',
    'about', 'up', 'out', 'off', 'over', 'under', 'again', 'further',
    'each', 'every', 'all', 'any', 'both', 'few', 'more', 'most',
    'other', 'some', 'such', 'only', 'own', 'same', 'between',
    'after', 'before', 'during', 'while', 'still', 'already',
    'use', 'used', 'uses', 'way', 'ways', 'step', 'steps',
    'guide', 'tutorial', 'example', 'examples', 'best', 'practice',
    'practices', 'tip', 'tips', 'trick', 'tricks', 'method', 'methods',
    'new', 'old', 'one', 'two', 'multiple', 'several', 'different',
    'within', 'without', 'like', 'need', 'want', 'try', 'able',
    'easily', 'quickly', 'simple', 'simply', 'basic', 'advanced',
    'properly', 'correctly', 'effectively', 'efficiently',
}

# Action verb synonyms — canonical form for semantic matching
VERB_SYNONYMS = {
    'create': 'create', 'make': 'create', 'build': 'create',
    'set up': 'create', 'setup': 'create', 'generate': 'create',
    'compose': 'create', 'draft': 'create', 'design': 'create',
    'establish': 'create', 'initialize': 'create', 'init': 'create',
    'provision': 'create', 'spin up': 'create', 'instantiate': 'create',
    'define': 'create', 'construct': 'create', 'craft': 'create',
    'add': 'add', 'insert': 'add', 'include': 'add', 'put': 'add',
    'append': 'add', 'inject': 'add', 'embed': 'add', 'place': 'add',
    'imbed': 'add', 'stick': 'add', 'attach': 'add',
    'delete': 'remove', 'remove': 'remove', 'clear': 'remove',
    'erase': 'remove', 'destroy': 'remove', 'purge': 'remove',
    'discard': 'remove', 'drop': 'remove', 'wipe': 'remove',
    'uninstall': 'remove', 'eliminate': 'remove',
    'get rid of': 'remove', 'take down': 'remove', 'take out': 'remove',
    'take off': 'remove',
    'edit': 'change', 'modify': 'change', 'change': 'change',
    'update': 'change', 'alter': 'change', 'revise': 'change',
    'amend': 'change', 'tweak': 'change', 'adjust': 'change',
    'rename': 'change', 'customize': 'change', 'turn': 'change',
    'switch': 'change',
    'view': 'view', 'see': 'view', 'check': 'view',
    'look': 'view', 'show': 'show', 'display': 'show', 'get': 'view',
    'access': 'access', 'browse': 'view', 'inspect': 'view',
    'retrieve': 'view', 'fetch': 'view', 'pull': 'view',
    'read': 'view', 'list': 'view', 'locate': 'find',
    'preview': 'view', 'review': 'view', 'lookup': 'find',
    'look up': 'find', 'navigate': 'view', 'go to': 'view',
    'find': 'find', 'search': 'find',
    'export': 'export', 'download': 'download', 'extract': 'export',
    'save as': 'export', 'dump': 'export', 'backup': 'export',
    'back up': 'export', 'save': 'download',
    'import': 'import', 'upload': 'import', 'load': 'import',
    'ingest': 'import', 'bring in': 'import', 'pull in': 'import',
    'bulk upload': 'import', 'bulk import': 'import', 'restore': 'import',
    'move': 'move', 'transfer': 'move', 'migrate': 'move',
    'shift': 'move', 'relocate': 'move', 'drag': 'move',
    'reorder': 'move', 'rearrange': 'move', 'reorganize': 'move',
    'copy': 'copy', 'clone': 'copy', 'duplicate': 'copy',
    'replicate': 'copy', 'mirror': 'copy',
    'assign': 'assign', 'allocate': 'assign', 'delegate': 'assign',
    'reassign': 'assign', 'distribute': 'assign',
    'link': 'link', 'connect': 'link', 'associate': 'link',
    'relate': 'link', 'map': 'link', 'bind': 'link',
    'hyperlink': 'link',
    'configure': 'configure', 'personalize': 'configure',
    'tailor': 'configure', 'set': 'configure', 'specify': 'configure',
    'tune': 'configure',
    'manage': 'manage', 'organize': 'manage', 'handle': 'manage',
    'administer': 'manage', 'maintain': 'manage', 'govern': 'manage',
    'oversee': 'manage', 'control': 'manage',
    'track': 'track', 'monitor': 'track', 'log': 'track',
    'measure': 'track', 'audit': 'track', 'record': 'record',
    'observe': 'track', 'watch': 'track',
    'enable': 'enable', 'activate': 'enable', 'turn on': 'enable',
    'switch on': 'enable', 'allow': 'enable', 'unlock': 'unlock',
    'disable': 'disable', 'deactivate': 'disable', 'turn off': 'disable',
    'switch off': 'disable', 'block': 'disable', 'restrict': 'disable',
    'mute': 'disable',
    'hide': 'hide', 'suppress': 'hide', 'conceal': 'hide',
    'reveal': 'show', 'unhide': 'show',
    'filter': 'filter', 'sort': 'sort', 'query': 'filter',
    'narrow': 'filter', 'refine': 'filter',
    'close': 'close', 'complete': 'close', 'finish': 'close',
    'resolve': 'close', 'archive': 'close', 'end': 'close',
    'start': 'start', 'begin': 'start', 'open': 'open',
    'launch': 'start', 'initiate': 'start', 'trigger': 'start',
    'run': 'run', 'execute': 'run', 'play': 'run',
    'stop': 'stop', 'pause': 'stop',
    'share': 'share', 'send': 'share', 'publish': 'share',
    'post': 'share', 'broadcast': 'share', 'forward': 'share',
    'submit': 'share',
    'approve': 'approve', 'accept': 'approve', 'confirm': 'approve',
    'authorize': 'approve', 'validate': 'approve',
    'deny': 'deny', 'reject': 'deny', 'decline': 'deny', 'refuse': 'deny',
    'merge': 'merge', 'combine': 'merge', 'consolidate': 'merge',
    'unify': 'merge', 'join': 'merge', 'aggregate': 'merge',
    'split': 'split', 'separate': 'split', 'divide': 'split',
    'decompose': 'split',
    'schedule': 'schedule', 'plan': 'schedule', 'book': 'schedule',
    'reserve': 'schedule',
    'automate': 'automate', 'batch': 'automate', 'bulk': 'automate',
    'sync': 'sync', 'synchronize': 'sync', 'synch': 'sync',
    'integrate': 'integrate', 'integration': 'integrate',
    'fix': 'fix', 'repair': 'fix', 'troubleshoot': 'fix',
    'debug': 'fix', 'diagnose': 'fix', 'solve': 'fix',
    'workaround': 'fix', 'patch': 'fix', 'recover': 'fix',
    'convert': 'convert', 'transform': 'convert', 'translate': 'translate',
    'format': 'format', 'parse': 'convert',
    'resize': 'resize', 'scale': 'resize',
    'crop': 'crop', 'trim': 'crop',
    'rotate': 'rotate',
    'flip': 'flip',
    'align': 'align', 'center': 'align',
    'animate': 'animate',
    'highlight': 'highlight',
    'group': 'group', 'ungroup': 'ungroup',
    'print': 'print',
    'wrap': 'wrap',
    'cite': 'cite', 'reference': 'cite',
    'compress': 'compress', 'condense': 'compress', 'shrink': 'compress',
    'minimize': 'compress', 'reduce': 'compress',
    'write': 'write', 'type': 'write',
    'draw': 'draw',
    'indent': 'indent',
    'outline': 'outline',
    'install': 'install',
    'use': 'use', 'apply': 'apply',
    'undo': 'undo',
    'collaborate': 'collaborate',
    'annotate': 'annotate',
    'zoom': 'zoom',
    'blur': 'blur',
    'cancel': 'cancel',
    'sign': 'sign',
    'lock': 'lock', 'freeze': 'lock',
    'narrate': 'record', 'capture': 'record',
    'loop': 'loop', 'autoplay': 'loop', 'repeat': 'loop',
    'present': 'present',
    'number': 'number',
    'get out of': 'exit', 'exit': 'exit',
    'get into': 'enter', 'enter': 'enter',
}

# Verb pairs that must NEVER merge (opposite intents)
NEVER_MERGE_VERB_PAIRS = {
    frozenset({"add", "remove"}),
    frozenset({"show", "hide"}),
    frozenset({"enable", "disable"}),
    frozenset({"lock", "unlock"}),
    frozenset({"group", "ungroup"}),
    frozenset({"import", "export"}),
    frozenset({"download", "upload"}),
    frozenset({"enter", "exit"}),
    frozenset({"run", "stop"}),
    frozenset({"flip", "rotate"}),
    frozenset({"record", "add"}),
    frozenset({"add", "create"}),
    frozenset({"open", "close"}),
    frozenset({"start", "stop"}),
}

OBJECT_SYNONYMS = {
    'ticket': 'issue', 'issue': 'issue', 'task': 'task',
    'item': 'issue', 'card': 'issue', 'work item': 'issue',
    'request': 'issue', 'incident': 'issue', 'case': 'issue',
    'bug': 'bug', 'defect': 'bug',
    'story': 'story', 'user story': 'story', 'requirement': 'story',
    'sprint': 'sprint', 'iteration': 'sprint', 'cycle': 'sprint',
    'board': 'board', 'kanban': 'board', 'scrum board': 'board',
    'kanban board': 'board', 'agile board': 'board',
    'backlog': 'backlog', 'product backlog': 'backlog',
    'epic': 'epic', 'initiative': 'epic',
    'project': 'project', 'workspace': 'workspace', 'space': 'workspace',
    'repo': 'repository', 'repository': 'repository',
    'folder': 'folder', 'directory': 'folder', 'collection': 'folder',
    'workflow': 'workflow', 'pipeline': 'pipeline', 'process': 'workflow',
    'flow': 'workflow',
    'automation': 'automation', 'rule': 'automation',
    'status': 'status', 'state': 'status',
    'dashboard': 'dashboard', 'overview': 'dashboard',
    'report': 'report', 'chart': 'report', 'graph': 'report',
    'analytics': 'report', 'metrics': 'report',
    'widget': 'widget', 'gadget': 'widget',
    'roadmap': 'roadmap', 'timeline': 'timeline', 'gantt': 'gantt',
    'field': 'field', 'custom field': 'field', 'attribute': 'field',
    'property': 'field', 'column': 'column',
    'label': 'label', 'tag': 'label',
    'priority': 'priority', 'severity': 'priority',
    'notification': 'notification', 'alert': 'notification',
    'reminder': 'notification',
    'comment': 'comment', 'note': 'comment', 'annotation': 'comment',
    'message': 'message', 'chat': 'message', 'dm': 'message',
    'subtask': 'subtask', 'sub-task': 'subtask', 'child issue': 'subtask',
    'child task': 'subtask',
    'attachment': 'attachment', 'file': 'file', 'document': 'document',
    'image': 'image', 'screenshot': 'image',
    'version': 'version', 'release': 'release',
    'permission': 'permission', 'role': 'role', 'privilege': 'permission',
    'user': 'user', 'member': 'user', 'account': 'account',
    'profile': 'profile', 'admin': 'admin',
    'template': 'template', 'blueprint': 'template', 'preset': 'template',
    'form': 'form', 'layout': 'layout', 'view': 'view', 'page': 'page',
    'api': 'api', 'rest api': 'api', 'endpoint': 'api',
    'webhook': 'webhook', 'callback': 'webhook',
    'plugin': 'plugin', 'add-on': 'plugin', 'addon': 'plugin',
    'extension': 'extension', 'app': 'app',
    'connector': 'integration', 'bridge': 'integration',
    'sla': 'sla', 'service level agreement': 'sla',
    'sso': 'sso', 'single sign on': 'sso', 'single sign-on': 'sso',
    'subscription': 'subscription', 'plan': 'plan',
    'payment': 'payment', 'billing': 'billing', 'invoice': 'invoice',
    'spreadsheet': 'spreadsheet', 'excel file': 'spreadsheet',
    'presentation': 'presentation', 'slide': 'slide', 'slides': 'slide',
    'video': 'video', 'audio': 'audio', 'recording': 'recording',
    'background': 'background', 'wallpaper': 'background',
    'font': 'font', 'text': 'text', 'paragraph': 'paragraph',
    'header': 'header', 'footer': 'footer',
    'table': 'table', 'grid': 'table',
    'chart': 'chart', 'diagram': 'diagram',
    'transition': 'transition', 'animation': 'animation',
    'theme': 'theme', 'color scheme': 'theme',
}


# ── XLSX Loading ─────────────────────────────────────────────────────────────

def parse_xlsx_tabs(xlsx_path):
    """Parse XLSX file, return dict of {tool_name: {type: 'standard'|'low_volume', keywords: [str]}}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    tools = {}

    for sheet_name in wb.sheetnames:
        name_lower = sheet_name.strip().lower()

        if any(skip in name_lower for skip in SKIP_TABS):
            continue
        if name_lower in ('dump', 'clean'):
            continue

        ws = wb[sheet_name]
        keywords = _extract_keywords_from_sheet(ws)

        if not keywords:
            continue

        is_low_vol = 'below 10' in name_lower or '< 10' in name_lower or '<10' in name_lower

        tool_name = _extract_tool_name(sheet_name)

        # Merge case-insensitive duplicates: find existing key with same lowercase
        existing_key = None
        for k in tools:
            if k.lower() == tool_name.lower():
                existing_key = k
                break
        if existing_key:
            tool_name = existing_key
        elif tool_name not in tools:
            tools[tool_name] = {'standard': [], 'low_volume': []}

        if is_low_vol:
            tools[tool_name]['low_volume'].extend(keywords)
        else:
            tools[tool_name]['standard'].extend(keywords)

    wb.close()
    return tools


def _extract_keywords_from_sheet(ws):
    """Extract keyword strings from a worksheet. Handles multi-column layouts."""
    keywords = []
    skip_values = {'name', 'keyword', 'keywords', 'query', 'term', 'terms',
                   'topic', 'title', 'url', '', 'none', 'pass 1', 'pass 2',
                   'pass 3', 'covered', 'done', 'done ', 'p1', 'p2', 'p3', 'existing',
                   'needs 1 final sonnet pass'}

    for row in ws.iter_rows(values_only=True):
        for cell_val in row:
            if cell_val is None:
                continue
            val = str(cell_val).strip()
            val_lower = val.lower()

            if val_lower in skip_values:
                continue
            if len(val) < 3:
                continue
            if val.startswith('ANTHROPIC_API_KEY'):
                continue
            if val.startswith('<loc>'):
                continue
            if val.startswith('http'):
                continue
            if val.startswith('**') and val.endswith('**'):
                continue

            if _looks_like_keyword(val):
                keywords.append(val)

    return keywords


def _looks_like_keyword(val):
    """Heuristic: is this cell value a keyword (vs metadata/command/noise)?"""
    if len(val) > 200:
        return False
    if val.startswith(('python3 ', 'ANTHROPIC_', '=')):
        return False
    if re.match(r'^[\d.]+$', val):
        return False
    if val.count('/') > 3:
        return False
    alpha_chars = sum(1 for c in val if c.isalpha())
    return alpha_chars >= 3


def _extract_tool_name(sheet_name):
    """Extract the tool name from a sheet tab name."""
    name = sheet_name.strip()
    name = re.sub(r'\s*-\s*(10\+|below 10|less than 10|<\s*10).*$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s*\(.*$', '', name)
    name = re.sub(r'\s*10\+\s*$', '', name, flags=re.IGNORECASE)
    return name.strip()


# ── CMS Loading ──────────────────────────────────────────────────────────────

def load_cms_titles(cms_path):
    """Load tutorial titles from CMS export CSV. Returns list of title strings."""
    titles = []
    with open(cms_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row.get('Name', '').strip()
            if name:
                titles.append(name)
    print(f"  Loaded {len(titles)} published tutorials from CMS")
    return titles


def build_cms_intent_index(titles):
    """Normalize CMS titles into intent keys for fuzzy matching.
    Returns dict of {intent_key: [original_titles]}.
    """
    index = defaultdict(list)
    for title in titles:
        key = _title_to_intent_key(title)
        if key:
            index[key].append(title)
    print(f"  Built CMS intent index: {len(index)} unique intents")
    return index


def _title_to_intent_key(title):
    """Convert a tutorial title like 'How to Add Columns in Notion' to an intent key."""
    text = title.lower().strip()

    for prefix in ['how to ', 'how do i ', 'how do you ', 'how can i ',
                    'how can you ', 'how can we ']:
        if text.startswith(prefix):
            text = text[len(prefix):]
            break

    text = re.sub(r'\b(a|an|the|my|your|our|this|that)\b', ' ', text)
    text = text.rstrip('?').strip()
    text = ' '.join(text.split())

    words = text.split()
    normalized = []
    for w in words:
        if w in STOP_WORDS and w not in ('in', 'on', 'to', 'from', 'with'):
            continue
        canon = VERB_SYNONYMS.get(w, w)
        canon = OBJECT_SYNONYMS.get(canon, canon)
        canon = _strip_plural(canon)
        normalized.append(canon)

    return tuple(sorted(set(normalized))) if normalized else None


# ── Keyword Normalization ────────────────────────────────────────────────────

def normalize_keyword(kw, tool_name):
    """Normalize a raw keyword into an intent key for deduplication."""
    text = kw.lower().strip()

    tool_lower = tool_name.lower()
    tool_variants = _tool_name_variants(tool_lower)
    for variant in sorted(tool_variants, key=len, reverse=True):
        text = text.replace(variant, ' ')

    for prefix in ['how to ', 'how do i ', 'how do you ', 'how can i ',
                    'how can you ', 'how can we ']:
        if text.startswith(prefix):
            text = text[len(prefix):]
            break

    text = re.sub(r'\b(a|an|the|my|your|our|this|that)\b', ' ', text)
    text = text.rstrip('?').strip()

    for phrase, canon in sorted(VERB_SYNONYMS.items(), key=lambda x: -len(x[0])):
        if ' ' in phrase:
            text = text.replace(phrase, canon)

    words = text.split()
    words = [VERB_SYNONYMS.get(w, w) for w in words]

    text_rejoined = ' '.join(words)
    for phrase, canon in sorted(OBJECT_SYNONYMS.items(), key=lambda x: -len(x[0])):
        if ' ' in phrase:
            text_rejoined = text_rejoined.replace(phrase, canon)
    words = text_rejoined.split()
    words = [OBJECT_SYNONYMS.get(w, w) for w in words]

    words = [w for w in words if w not in STOP_WORDS and len(w) > 1]
    normalized = [_strip_plural(w) for w in words]

    return tuple(sorted(set(normalized)))


def _strip_plural(w):
    """Simple plural stripping."""
    if w.endswith('ies') and len(w) > 4:
        return w[:-3] + 'y'
    if w.endswith('ses') and len(w) > 4:
        return w[:-2]
    if w.endswith('s') and not w.endswith('ss') and not w.endswith('us') and len(w) > 3:
        return w[:-1]
    return w


def _tool_name_variants(tool_lower):
    """Generate variants of a tool name for removal during normalization."""
    variants = {tool_lower}
    if '.' in tool_lower:
        variants.add(tool_lower.replace('.', ''))
    parts = tool_lower.split()
    if len(parts) > 1:
        variants.add(''.join(parts))
    variants.add(f"{tool_lower} software")
    variants.add(f"{tool_lower} cloud")
    return variants


# ── Categorization ───────────────────────────────────────────────────────────

def categorize_keyword(kw, tool_name):
    """Categorize: integration, howto, tutorial, or other."""
    kw_lower = kw.lower()

    integration_signals = ['integration', 'integrate', 'connect ', 'connect to',
                           'connect with', 'sync ', 'synced', 'connector']
    if any(sig in kw_lower for sig in integration_signals):
        return 'integration'

    if kw_lower.startswith(('how to', 'how do', 'how can')):
        return 'howto'

    if 'tutorial' in kw_lower:
        return 'tutorial'

    return 'other'


def is_irrelevant(kw, tool_name):
    """Filter out non-tutorial keywords."""
    kw_lower = kw.lower().strip()
    tool_lower = tool_name.lower()

    exact_irrelevant = {
        f'{tool_lower} tutorial',
        f'how to use {tool_lower}',
        f'{tool_lower} how to',
        f'how to {tool_lower}',
        f'tutorial {tool_lower}',
        f'{tool_lower} tutorial for beginners',
        f'learn how to use {tool_lower}',
        f'best {tool_lower} tutorial',
    }
    if kw_lower in exact_irrelevant:
        return True

    irrelevant_patterns = [
        'how to pronounce', 'how to spell', 'how to say ',
        'tutorial pdf', 'tutorial video', 'tutorial youtube',
        'tutorial udemy', 'tutorial guru99', 'tutorial ppt',
        'tutorial hindi', 'tutorial arabic', 'tutorial deutsch',
        'tutorial español', 'tutorial français', 'tutorial pl',
        'tutorial for beginners video', 'tutorial for beginners pdf',
        'best tutorial', 'free tutorial', 'tutorial free',
        'tutorial 2018', 'tutorial 2019', 'tutorial 2020',
        'tutorial 2021', 'tutorial 2022', 'tutorial 2023', 'tutorial 2024',
        'how to learn ', 'certification', 'interview', 'resume',
        'what is ', 'what does ', 'does ', ' vs ',
        'how to contact', 'how to cancel subscription',
        'how to get a job', 'salary', 'hiring', 'career',
        'how to pronounce', 'pronunciation',
        'template', 'templates', 'free download',
        'alternative', 'alternatives', 'pricing', 'price', 'cost',
        'review', 'reviews', 'reddit', 'quora',
        'login', 'log in', 'sign in', 'sign up',
        'not working', 'error', 'issue', 'problem', 'bug',
        'announcement', 'release note', 'changelog',
        'coupon', 'discount', 'promo',
    ]
    if any(p in kw_lower for p in irrelevant_patterns):
        return True

    if _is_non_english(kw):
        return True

    return False


def _is_non_english(text):
    """Detect non-English keywords by checking for non-ASCII letter ratio."""
    letters = [c for c in text if c.isalpha()]
    if not letters:
        return False
    non_ascii = sum(1 for c in letters if ord(c) > 127)
    return non_ascii / len(letters) > 0.3


# ── Core Pipeline ────────────────────────────────────────────────────────────

def deduplicate_keywords(keywords, tool_name):
    """Group keywords by normalized intent key. Returns dict of {intent_key: [original_keywords]}."""
    groups = defaultdict(list)
    for kw in keywords:
        key = normalize_keyword(kw, tool_name)
        if key:
            groups[key].append(kw)
    return groups


def word_overlap_score(words_a, words_b):
    """Jaccard similarity between two word tuples."""
    if not words_a or not words_b:
        return 0.0
    set_a = set(words_a)
    set_b = set(words_b)
    intersection = set_a & set_b
    union = set_a | set_b
    return len(intersection) / len(union) if union else 0.0


def _get_lead_verb(intent_key):
    """Extract the likely lead verb from an intent key for safety checks."""
    canonical_verbs = set(VERB_SYNONYMS.values())
    for word in intent_key:
        if word in canonical_verbs:
            return word
    return None


def merge_similar_groups(groups, tool_name, threshold=0.70):
    """Second pass: merge groups with high word overlap using centroid clustering.
    Returns merged dict of {intent_key: [keywords]}.
    """
    items = list(groups.items())
    items.sort(key=lambda x: -len(x[1]))

    clusters = []  # (rep_key, [all_keywords])
    word_to_clusters = defaultdict(list)

    for key, kws in items:
        set_key = set(key)
        best_cluster = None
        best_score = 0.0

        candidate_clusters = set()
        for w in key:
            for c_idx in word_to_clusters.get(w, []):
                candidate_clusters.add(c_idx)

        for c_idx in candidate_clusters:
            rep_key, _ = clusters[c_idx]

            verb_a = _get_lead_verb(key)
            verb_b = _get_lead_verb(rep_key)
            if verb_a and verb_b and frozenset({verb_a, verb_b}) in NEVER_MERGE_VERB_PAIRS:
                continue

            score = word_overlap_score(key, rep_key)
            if score >= threshold and score > best_score:
                best_score = score
                best_cluster = c_idx

        if best_cluster is not None:
            clusters[best_cluster][1].extend(kws)
        else:
            c_idx = len(clusters)
            clusters.append((key, list(kws)))
            for w in key:
                word_to_clusters[w].append(c_idx)

    return {rep_key: kws for rep_key, kws in clusters}


def pick_canonical(variants, tool_name):
    """Pick the best representative keyword from a group of variants."""
    howtos = [v for v in variants if v.lower().startswith('how to')]
    if howtos:
        with_article = [v for v in howtos if ' a ' in v.lower() or ' an ' in v.lower()]
        if with_article:
            return min(with_article, key=len)
        return min(howtos, key=len)

    integrations = [v for v in variants if 'integration' in v.lower()]
    if integrations:
        return min(integrations, key=len)

    return min(variants, key=len)


def check_cms_coverage(intent_key, cms_index):
    """Check if an intent key matches any published tutorial."""
    if intent_key in cms_index:
        return True

    for cms_key in cms_index:
        if not cms_key or not intent_key:
            continue
        score = word_overlap_score(intent_key, cms_key)
        if score >= 0.80:
            return True

    return False


def check_cms_coverage_batch(merged_groups, cms_index):
    """Efficiently check all groups against CMS index.
    Uses inverted index for speed instead of O(n*m) comparison.
    """
    cms_word_index = defaultdict(set)
    for cms_key in cms_index:
        if cms_key:
            for w in cms_key:
                cms_word_index[w].add(cms_key)

    covered = set()
    for intent_key in merged_groups:
        if intent_key in cms_index:
            covered.add(intent_key)
            continue

        candidate_cms_keys = set()
        for w in intent_key:
            for ck in cms_word_index.get(w, []):
                candidate_cms_keys.add(ck)

        for cms_key in candidate_cms_keys:
            score = word_overlap_score(intent_key, cms_key)
            if score >= 0.80:
                covered.add(intent_key)
                break

    return covered


def process_tool(tool_name, standard_kws, low_vol_kws, cms_index, min_freq=6):
    """Full pipeline for one tool. Returns list of result dicts."""
    print(f"\n{'='*70}")
    print(f"  PROCESSING: {tool_name}")
    print(f"  Standard keywords: {len(standard_kws)}, Low-volume: {len(low_vol_kws)}")
    print(f"{'='*70}")

    results = []

    # ── Process standard (10+) keywords ──
    if standard_kws:
        filtered = [kw for kw in standard_kws if not is_irrelevant(kw, tool_name)]
        print(f"  Standard: {len(standard_kws)} raw → {len(filtered)} after irrelevant filter")

        groups = deduplicate_keywords(filtered, tool_name)
        print(f"  Standard: {len(filtered)} → {len(groups)} intent groups (pass 1)")

        merged = merge_similar_groups(groups, tool_name, threshold=0.70)
        print(f"  Standard: {len(groups)} → {len(merged)} after synonym merge (pass 2)")

        covered = check_cms_coverage_batch(merged, cms_index)
        print(f"  Standard: {len(covered)} already covered in CMS")

        for key, kws in sorted(merged.items(), key=lambda x: -len(x[1])):
            if key in covered:
                continue
            if len(kws) < 2:
                continue

            canonical = pick_canonical(kws, tool_name)
            cat = categorize_keyword(canonical, tool_name)
            signal = 'high' if len(kws) >= 5 else 'medium' if len(kws) >= 3 else 'low'

            results.append({
                'canonical_keyword': canonical,
                'category': cat,
                'variant_count': len(kws),
                'signal_strength': signal,
                'source': 'volume_10+',
                'frequency_count': '',
                'all_variants': ' | '.join(kws[:15]),
            })

    # ── Process low-volume (<10) keywords with frequency clustering ──
    if low_vol_kws:
        filtered_lv = [kw for kw in low_vol_kws if not is_irrelevant(kw, tool_name)]
        print(f"  Low-vol: {len(low_vol_kws)} raw → {len(filtered_lv)} after irrelevant filter")

        groups_lv = deduplicate_keywords(filtered_lv, tool_name)
        print(f"  Low-vol: {len(filtered_lv)} → {len(groups_lv)} intent groups (pass 1)")

        merged_lv = merge_similar_groups(groups_lv, tool_name, threshold=0.65)
        print(f"  Low-vol: {len(groups_lv)} → {len(merged_lv)} after aggressive merge (pass 2)")

        covered_lv = check_cms_coverage_batch(merged_lv, cms_index)
        print(f"  Low-vol: {len(covered_lv)} already covered in CMS")

        freq_qualified = 0
        for key, kws in sorted(merged_lv.items(), key=lambda x: -len(x[1])):
            if key in covered_lv:
                continue
            if len(kws) < min_freq:
                continue

            canonical = pick_canonical(kws, tool_name)
            cat = categorize_keyword(canonical, tool_name)

            results.append({
                'canonical_keyword': canonical,
                'category': cat,
                'variant_count': len(kws),
                'signal_strength': 'frequency',
                'source': 'frequency_matched',
                'frequency_count': str(len(kws)),
                'all_variants': ' | '.join(kws[:15]),
            })
            freq_qualified += 1

        print(f"  Low-vol: {freq_qualified} keywords qualified (frequency >= {min_freq})")

    # ── Cross-batch dedup between standard and frequency results ──
    if results:
        results = _cross_dedup_results(results, tool_name)

    results.sort(key=lambda x: (-x['variant_count'], x['category']))

    std_count = sum(1 for r in results if r['source'] == 'volume_10+')
    freq_count = sum(1 for r in results if r['source'] == 'frequency_matched')
    print(f"\n  TOTAL: {len(results)} keywords ({std_count} volume-based, {freq_count} frequency-matched)")

    return results


def _cross_dedup_results(results, tool_name):
    """Remove frequency-matched results that duplicate volume-based ones."""
    vol_keys = set()
    for r in results:
        if r['source'] == 'volume_10+':
            key = normalize_keyword(r['canonical_keyword'], tool_name)
            vol_keys.add(key)

    deduped = []
    dropped = 0
    for r in results:
        if r['source'] == 'frequency_matched':
            key = normalize_keyword(r['canonical_keyword'], tool_name)
            is_dup = False
            for vk in vol_keys:
                if word_overlap_score(key, vk) >= 0.75:
                    is_dup = True
                    break
            if is_dup:
                dropped += 1
                continue
        deduped.append(r)

    if dropped:
        print(f"  Cross-dedup: removed {dropped} frequency keywords that overlap with volume ones")
    return deduped


# ── Multi-tool tab handling ──────────────────────────────────────────────────

def parse_mix_of_tools_tab(ws):
    """Parse the 'mix of tools' tab where each column is a separate tool."""
    tools = {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return tools

    header = rows[0]
    for col_idx, tool_name in enumerate(header):
        if tool_name is None:
            continue
        tool_name = str(tool_name).strip()
        if not tool_name or tool_name.lower() in ('none', ''):
            continue

        keywords = []
        for row in rows[1:]:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip()
                if val and _looks_like_keyword(val) and len(val) >= 3:
                    keywords.append(val)

        if keywords:
            tools[tool_name] = {'standard': keywords, 'low_volume': []}

    return tools


# ── Output ───────────────────────────────────────────────────────────────────

def write_tool_csv(tool_name, results, output_dir):
    """Write results for one tool to CSV."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[^\w\s.-]', '', tool_name).strip().replace(' ', '_')
    path = output_dir / f"{safe_name}_cleaned.csv"

    fieldnames = [
        'canonical_keyword', 'suggested_title', 'category',
        'variant_count', 'signal_strength', 'source',
        'frequency_count', 'all_variants',
    ]

    kept = 0
    dropped = 0
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in results:
            row = dict(r)
            title = _generate_title(r, tool_name)
            if title is None:
                dropped += 1
                continue
            row['suggested_title'] = title
            writer.writerow(row)
            kept += 1

    if dropped:
        print(f"  Quality gate: dropped {dropped}, kept {kept}")
    return path


def _generate_title(result, tool_name):
    """Convert any keyword into a proper 'How to X in [Tool]' tutorial title."""
    canonical = result['canonical_keyword']
    cat = result['category']
    tool_cap = _capitalize_tool(tool_name)

    # Already a how-to — just clean it up
    if canonical.lower().startswith(('how to ', 'how do ', 'how can ')):
        title = canonical
        title = re.sub(re.escape(tool_name), tool_cap, title, flags=re.IGNORECASE)
        words = title.split()
        if words:
            words[0] = 'How'
            if len(words) > 1:
                words[1] = words[1].lower()
        title = ' '.join(words)
        title = _fix_known_tool_names(title)
        if tool_cap.lower() not in title.lower():
            title += f" in {tool_cap}"
        return _clean_title(title)

    # Integration → "How to Integrate [Tool] with [Partner]"
    if cat == 'integration':
        partner = _extract_partner(canonical, tool_name)
        if partner:
            partner_cap = _capitalize_tool(partner)
            title = f"How to Integrate {tool_cap} with {partner_cap}"
            return _clean_title(_fix_known_tool_names(title))
        return None

    # Other/tutorial — try to extract an action and object
    return _convert_to_howto(canonical, tool_name, tool_cap)


def _capitalize_tool(name):
    """Proper-case a tool name."""
    special = {
        'capcut': 'CapCut', 'docker': 'Docker', 'kubernetes': 'Kubernetes',
        'midjourney': 'Midjourney', 'terraform': 'Terraform', 'asana': 'Asana',
        'github': 'GitHub', 'gitlab': 'GitLab', 'hubspot': 'HubSpot',
        'salesforce': 'Salesforce', 'netsuite': 'NetSuite', 'jira': 'Jira',
        'slack': 'Slack', 'trello': 'Trello', 'notion': 'Notion',
        'airtable': 'Airtable', 'clickup': 'ClickUp', 'monday.com': 'Monday.com',
        'zendesk': 'Zendesk', 'servicenow': 'ServiceNow', 'confluence': 'Confluence',
        'bitbucket': 'Bitbucket', 'datadog': 'Datadog', 'pagerduty': 'PagerDuty',
        'figma': 'Figma', 'zapier': 'Zapier', 'miro': 'Miro',
        'power bi': 'Power BI', 'sharepoint': 'SharePoint', 'outlook': 'Outlook',
        'gmail': 'Gmail', 'google drive': 'Google Drive', 'dropbox': 'Dropbox',
        'postgresql': 'PostgreSQL', 'mysql': 'MySQL', 'mongodb': 'MongoDB',
        'redis': 'Redis', 'nginx': 'NGINX', 'jenkins': 'Jenkins',
        'ansible': 'Ansible', 'prometheus': 'Prometheus', 'grafana': 'Grafana',
        'elasticsearch': 'Elasticsearch', 'aws': 'AWS', 'azure': 'Azure',
        'gcp': 'GCP', 'heroku': 'Heroku', 'vercel': 'Vercel',
        'tiktok': 'TikTok', 'instagram': 'Instagram', 'youtube': 'YouTube',
        'linkedin': 'LinkedIn', 'facebook': 'Facebook', 'whatsapp': 'WhatsApp',
        'discord': 'Discord', 'reddit': 'Reddit',
        'quickbooks': 'QuickBooks', 'xero': 'Xero', 'stripe': 'Stripe',
        'shopify': 'Shopify', 'woocommerce': 'WooCommerce',
        'mailchimp': 'Mailchimp', 'sendgrid': 'SendGrid',
        'okta': 'Okta', 'auth0': 'Auth0',
        'n8n': 'n8n', 'ollama': 'Ollama',
    }
    return special.get(name.lower(), name.title())


def _clean_title(title):
    """Final cleanup: remove trailing punctuation, fix spacing."""
    title = title.rstrip('?.!,;:').strip()
    title = re.sub(r'\s+', ' ', title)
    return title


def _fix_known_tool_names(title):
    """Fix casing of known tool/product names wherever they appear in a title."""
    known = {
        'jira': 'Jira', 'slack': 'Slack', 'trello': 'Trello', 'notion': 'Notion',
        'asana': 'Asana', 'github': 'GitHub', 'gitlab': 'GitLab',
        'hubspot': 'HubSpot', 'salesforce': 'Salesforce', 'netsuite': 'NetSuite',
        'zendesk': 'Zendesk', 'confluence': 'Confluence', 'bitbucket': 'Bitbucket',
        'servicenow': 'ServiceNow', 'monday.com': 'Monday.com',
        'clickup': 'ClickUp', 'airtable': 'Airtable', 'figma': 'Figma',
        'zapier': 'Zapier', 'miro': 'Miro', 'docker': 'Docker',
        'kubernetes': 'Kubernetes', 'terraform': 'Terraform',
        'capcut': 'CapCut', 'midjourney': 'Midjourney',
        'google calendar': 'Google Calendar', 'google drive': 'Google Drive',
        'google sheets': 'Google Sheets', 'google docs': 'Google Docs',
        'gmail': 'Gmail', 'outlook': 'Outlook', 'teams': 'Teams',
        'sharepoint': 'SharePoint', 'excel': 'Excel',
        'power bi': 'Power BI', 'power automate': 'Power Automate',
        'tiktok': 'TikTok', 'instagram': 'Instagram', 'youtube': 'YouTube',
        'linkedin': 'LinkedIn', 'facebook': 'Facebook', 'whatsapp': 'WhatsApp',
        'discord': 'Discord', 'reddit': 'Reddit',
        'postgresql': 'PostgreSQL', 'mysql': 'MySQL', 'mongodb': 'MongoDB',
        'redis': 'Redis', 'nginx': 'NGINX', 'jenkins': 'Jenkins',
        'ansible': 'Ansible', 'prometheus': 'Prometheus', 'grafana': 'Grafana',
        'aws': 'AWS', 'azure': 'Azure', 'gcp': 'GCP',
        'quickbooks': 'QuickBooks', 'xero': 'Xero', 'stripe': 'Stripe',
        'shopify': 'Shopify', 'mailchimp': 'Mailchimp',
        'okta': 'Okta', 'datadog': 'Datadog', 'pagerduty': 'PagerDuty',
        'todoist': 'Todoist', 'dropbox': 'Dropbox', 'ubuntu': 'Ubuntu',
        'centos': 'CentOS', 'ollama': 'Ollama', 'n8n': 'n8n',
        'chatgpt': 'ChatGPT', 'openai': 'OpenAI',
        'smartsheet': 'Smartsheet', 'clockify': 'Clockify',
        'wrike': 'Wrike', 'basecamp': 'Basecamp', 'evernote': 'Evernote',
        'harvest': 'Harvest', 'pipedrive': 'Pipedrive', 'zoho': 'Zoho',
        'freshdesk': 'Freshdesk', 'freshservice': 'Freshservice',
        'intercom': 'Intercom', 'hootsuite': 'Hootsuite',
        'microsoft': 'Microsoft', 'claude': 'Claude', 'copilot': 'Copilot',
    }
    for lower_name, proper in sorted(known.items(), key=lambda x: -len(x[0])):
        title = re.sub(r'\b' + re.escape(lower_name) + r'\b', proper, title, flags=re.IGNORECASE)
    return title


def _extract_partner(kw, tool_name):
    """Extract the integration partner name from a keyword."""
    kw_lower = kw.lower()
    tool_lower = tool_name.lower()
    tool_variants = _tool_name_variants(tool_lower)

    cleaned = kw_lower
    for variant in sorted(tool_variants, key=len, reverse=True):
        cleaned = re.sub(r'\b' + re.escape(variant) + r'\b', ' ', cleaned)

    # Use word-boundary removal to avoid mangling tool names
    remove_words = [
        'integration', 'integrate', 'integrated', 'integrating',
        'connect', 'connected', 'connecting', 'connection', 'connector',
        'sync', 'synced', 'syncing', 'synchronize',
        'how to', 'how do i', 'how can i',
        'with', 'to', 'from', 'and', 'for', 'the', 'an', 'in', 'on',
        'using', 'via', 'through', 'between', 'not working', 'best', 'free',
        'official', 'documentation', 'guide', 'setup', 'steps',
        'webhook', 'rest api', 'api', 'plugin', 'app', 'marketplace',
        'cannot', "can't",
    ]
    for word in sorted(remove_words, key=len, reverse=True):
        cleaned = re.sub(r'\b' + re.escape(word) + r'\b', ' ', cleaned)

    cleaned = ' '.join(cleaned.split()).strip()

    words = [w.strip().strip('?.,!') for w in cleaned.split() if w.strip()]
    words = [w for w in words if len(w) > 1 and w not in {'of', 'at', 'by', 'if', 'is', 'or', 'a'}]

    partner = ' '.join(words).strip()
    return partner if partner else None


def _convert_to_howto(kw, tool_name, tool_cap):
    """Attempt to convert a raw keyword into how-to format."""
    kw_lower = kw.lower().strip()
    tool_lower = tool_name.lower()

    # Remove tool name
    cleaned = kw_lower
    for variant in sorted(_tool_name_variants(tool_lower), key=len, reverse=True):
        cleaned = cleaned.replace(variant, ' ')
    cleaned = ' '.join(cleaned.split()).strip()

    # Remove generic noise words
    for noise in ['tutorial', 'guide', 'step by step', 'for beginners',
                  'basics', 'advanced', 'complete', 'full']:
        cleaned = cleaned.replace(noise, ' ')
    cleaned = ' '.join(cleaned.split()).strip()

    if not cleaned or len(cleaned) < 3:
        return None

    # Try to identify an action verb at the start
    words = cleaned.split()
    canonical_verbs = set(VERB_SYNONYMS.values())
    all_verb_forms = set(VERB_SYNONYMS.keys()) | canonical_verbs

    if words[0] in all_verb_forms:
        action = words[0]
        obj = ' '.join(words[1:]) if len(words) > 1 else ''
        if obj:
            title = f"How to {action.title()} {obj.title()} in {tool_cap}"
        else:
            title = f"How to {action.title()} in {tool_cap}"
        return _fix_known_tool_names(title)

    # No action verb found — treat the whole thing as a noun/object
    title = f"How to Use {cleaned.title()} in {tool_cap}"
    return _fix_known_tool_names(title)


# ── AI Cleanup (optional) ───────────────────────────────────────────────────

def ai_cleanup(results, tool_name, api_key, batch_size=25):
    """Optional AI pass to fix titles and catch remaining duplicates."""
    try:
        import anthropic
    except ImportError:
        print("  WARNING: anthropic package not installed, skipping AI cleanup")
        return results

    client = anthropic.Anthropic(api_key=api_key)
    model = "claude-haiku-4-5-20251001"

    total_batches = (len(results) + batch_size - 1) // batch_size
    print(f"\n  AI cleanup: {len(results)} entries in {total_batches} batches (model: {model})")

    all_fixes = {}
    duplicate_map = {}

    for batch_num in range(total_batches):
        start = batch_num * batch_size
        end = min(start + batch_size, len(results))
        batch = results[start:end]

        print(f"    Batch {batch_num + 1}/{total_batches}...", end=" ", flush=True)

        entries_text = ""
        for i, entry in enumerate(batch):
            entries_text += f"\n[{i}] keyword: {entry['canonical_keyword']}"
            entries_text += f"\n    category: {entry['category']}"
            entries_text += f"\n    variants: {entry['all_variants'][:200]}"
            entries_text += "\n"

        prompt = f"""Clean up this keyword dataset for {tool_name} tutorials.

For each entry:
1. Generate a proper tutorial title (e.g., "How to [Action] in {tool_name}")
2. Fix any garbled names using the variants as clues
3. Flag if irrelevant (not tutorial-worthy)
4. Flag if duplicate of another entry in this batch (by index)

Entries:
{entries_text}

Return ONLY a JSON array:
[{{"index": 0, "fixed_title": "...", "irrelevant": false, "duplicate_of": null}}, ...]"""

        try:
            response = client.messages.create(
                model=model,
                max_tokens=4096,
                messages=[{"role": "user", "content": prompt}],
            )
            text = response.content[0].text.strip()
            if text.startswith("```"):
                text = text.split("\n", 1)[1]
                if text.endswith("```"):
                    text = text[:-3]
                text = text.strip()

            fixes = json.loads(text)
            for fix in fixes:
                local_idx = fix['index']
                global_idx = start + local_idx
                if global_idx < len(results):
                    all_fixes[global_idx] = fix
                    if fix.get('duplicate_of') is not None:
                        duplicate_map[global_idx] = start + fix['duplicate_of']

            print(f"OK")
        except Exception as e:
            print(f"FAILED ({e})")

        if batch_num < total_batches - 1:
            time.sleep(0.5)

    # Apply fixes
    fixed_count = 0
    irrelevant_indices = set()
    for idx, fix in all_fixes.items():
        if fix.get('irrelevant'):
            irrelevant_indices.add(idx)
        if fix.get('fixed_title'):
            results[idx]['_ai_title'] = fix['fixed_title']
            fixed_count += 1

    # Remove irrelevant and duplicates
    skip = irrelevant_indices | set(duplicate_map.keys())
    final = [r for i, r in enumerate(results) if i not in skip]

    for r in final:
        if '_ai_title' in r:
            r['canonical_keyword'] = r.pop('_ai_title')
        else:
            r.pop('_ai_title', None)

    print(f"  AI cleanup: {fixed_count} titles fixed, {len(irrelevant_indices)} irrelevant, "
          f"{len(duplicate_map)} duplicates → {len(final)} final")
    return final


# ── Global cross-tool dedup ──────────────────────────────────────────────────

def global_dedup_across_tools(all_results):
    """Check for duplicate keywords that appear across different tool outputs.
    This is informational — it adds a flag but doesn't remove entries.
    """
    seen = {}
    for tool, results in all_results.items():
        for r in results:
            kw_lower = r['canonical_keyword'].lower().strip()
            if kw_lower in seen and seen[kw_lower] != tool:
                r['cross_tool_dup'] = seen[kw_lower]
            else:
                seen[kw_lower] = tool


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Keyword Consolidation Pipeline')
    parser.add_argument('--xlsx', required=True, help='Input XLSX file with keyword tabs')
    parser.add_argument('--cms', required=True, help='CMS export CSV (Webflow tutorials)')
    parser.add_argument('--tools', help='Comma-separated list of tools to process (default: all)')
    parser.add_argument('--output', '-o', default='output', help='Output directory (default: output/)')
    parser.add_argument('--min-freq', type=int, default=6,
                        help='Min frequency for low-volume keywords (default: 6)')
    parser.add_argument('--threshold', type=float, default=0.70,
                        help='Word overlap threshold for merging (default: 0.70)')
    parser.add_argument('--ai-cleanup', action='store_true',
                        help='Run optional AI cleanup pass (requires ANTHROPIC_API_KEY)')

    args = parser.parse_args()

    print(f"{'='*70}")
    print(f"  KEYWORD CONSOLIDATION PIPELINE")
    print(f"{'='*70}")

    # Load CMS
    print(f"\nLoading CMS export...")
    cms_titles = load_cms_titles(args.cms)
    cms_index = build_cms_intent_index(cms_titles)

    # Parse XLSX
    print(f"\nParsing XLSX tabs...")
    tools = parse_xlsx_tabs(args.xlsx)

    # Handle 'mix of tools' tab separately
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() == 'mix of tools':
            ws = wb[sheet_name]
            mix_tools = parse_mix_of_tools_tab(ws)
            for t, kws in mix_tools.items():
                if t not in tools:
                    tools[t] = {'standard': [], 'low_volume': []}
                tools[t]['standard'].extend(kws['standard'])
    wb.close()

    print(f"\nFound {len(tools)} tools:")
    for t in sorted(tools.keys()):
        std = len(tools[t]['standard'])
        lv = len(tools[t]['low_volume'])
        print(f"  {t}: {std} standard + {lv} low-volume")

    # Filter to requested tools
    if args.tools:
        requested = {t.strip().lower() for t in args.tools.split(',')}
        tools = {t: kws for t, kws in tools.items() if t.lower() in requested}
        print(f"\nFiltered to {len(tools)} requested tools")

    # API key for AI cleanup
    api_key = None
    if args.ai_cleanup:
        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key and Path('.api_key').exists():
            api_key = Path('.api_key').read_text().strip()
        if not api_key:
            print("WARNING: --ai-cleanup requested but no ANTHROPIC_API_KEY found")
            args.ai_cleanup = False

    # Process each tool
    all_results = {}
    for tool_name in sorted(tools.keys()):
        data = tools[tool_name]
        results = process_tool(
            tool_name,
            data['standard'],
            data['low_volume'],
            cms_index,
            min_freq=args.min_freq,
        )

        if args.ai_cleanup and results:
            results = ai_cleanup(results, tool_name, api_key)

        if results:
            path = write_tool_csv(tool_name, results, args.output)
            print(f"  Written: {path}")
            all_results[tool_name] = results

    # Summary
    print(f"\n{'='*70}")
    print(f"  SUMMARY")
    print(f"{'='*70}")
    total = 0
    for tool_name in sorted(all_results.keys()):
        r = all_results[tool_name]
        std = sum(1 for x in r if x['source'] == 'volume_10+')
        freq = sum(1 for x in r if x['source'] == 'frequency_matched')
        print(f"  {tool_name}: {len(r)} keywords ({std} vol + {freq} freq)")
        total += len(r)
    print(f"\n  TOTAL: {total} clean keywords across {len(all_results)} tools")
    print(f"  Output directory: {args.output}/\n")


if __name__ == '__main__':
    main()
